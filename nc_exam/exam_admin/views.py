#!/usr/bin/env python3
# -*- coding:utf-8 -*-
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.db import connection
from django.db.models import Q
#from exam_admin.models import Department, WorkType, ExamPapers, TestPapers, Members, Papers, PaperTypes, Position, Questions, PaperImportLog, PaperPositionRange, PaperWorkTypeRange, PaperDepRange
from exam_admin.models import *
from django.http import HttpResponse
from datetime import datetime
from datetime import timedelta
from django.db.models import Count, Sum, Max, Avg
from functools import reduce
from openpyxl import load_workbook, Workbook

import simplejson as json
import logging, os, sys, operator, uuid
# Create your views here.
def admin(request):
    return render(request, 'admin_main.html')

@csrf_exempt
def getTopsComboData(request):
    combo = request.POST.get('combo', None)
    if not combo:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % combo, content_type='text/json')

    ret_list = list()
    try:
        if combo == 'workshop':
            for dep in Department.objects.filter(level__lt=7).order_by('level'):
                ret_list.append({ 'value': dep.dep_id, 'label': dep.dep_name })
        elif combo == 'worktype':
            for wt in WorkType.objects.all().order_by('work_type_id'):
                ret_list.append({ 'value': wt.work_type_id, 'label': wt.type_name })
        elif combo == 'position':
            for pos in Position.objects.all().order_by('position_id'):
                ret_list.append({ 'value': pos.position_id, 'label': pos.name })

        return HttpResponse(json.dumps({ 'data': ret_list }), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"无法获取列表：%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getTopsComboDataByPaper(request):
    combo = request.POST.get('combo', None)
    paper_ids = request.POST.get('paper_ids', '')
    if not combo:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % combo, content_type='text/json')
    
    pid_arr = paper_ids.split(',')

    ret_list = list()
    try:
        if combo == 'workshop':
            if paper_ids != '':
                deps = PaperDepRange.objects.filter(paper_id__in=pid_arr).values_list('dep_id')
                deps_in_dep = Department.objects.filter(level__lt=7, dep_id__in=deps).order_by('level')
            else:
                deps_in_dep = Department.objects.all();

            for dep in deps_in_dep:
                ret_list.append({ 'value': dep.dep_id, 'label': dep.dep_name })
        elif combo == 'worktype':
            if paper_ids != '':
                wts = PaperWorkTypeRange.objects.filter(paper_id__in=pid_arr).values_list('work_type_id')
                wts_in_wt = WorkType.objects.filter(work_type_id__in=wts).order_by('work_type_id')
            else:
                wts_in_wt = WorkType.objects.all()

            for wt in wts_in_wt:
                ret_list.append({ 'value': wt.work_type_id, 'label': wt.type_name })
        elif combo == 'position':
            if paper_ids != '':
                poss = PaperPositionRange.objects.filter(paper_id__in=pid_arr).values_list('position_id')
                poss_in_pos = Position.objects.filter(position_id__in=poss).order_by('position_id')
            else:
                poss_in_pos = Position.objects.all()

            for pos in poss_in_pos:
                ret_list.append({ 'value': pos.position_id, 'label': pos.name })

        return HttpResponse(json.dumps({ 'data': ret_list }), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"无法获取列表：%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getTopsList(request):
    workshop = request.POST.get('workshop', None)
    worktype = request.POST.get('worktype', None)
    all_mem = request.POST.get('all', None)
    date_range = request.POST.get('date_range', None)

    if not workshop or not worktype or not all_mem or not date_range:
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    now = datetime.now()
    
    #本季
    month = (now.month - 1) - (now.month - 1) % 3 + 1
    this_quarter_start = datetime(now.year, month, 1)
    this_quarter_end = datetime(now.year, month + 3, 1) - timedelta(days=1)
    #本日
    this_day_start = datetime(now.year, month, now.day, hour=0, minute=0, second=0)
    this_day_end = datetime(now.year, month, now.day, hour=23, minute=59, second=59)
    #本周
    this_week_start = now - timedelta(days=now.weekday())
    this_week_end = now - timedelta(days=6-now.weekday())
    #本月
    this_month_start = datetime(now.year, now.month, 1)
    this_month_end = datetime(now.year, now.month + 1, 1) - timedelta(days=1)
    #本年
    this_year_start = datetime(now.year, 1, 1)
    this_year_end = datetime(now.year + 1, 1, 1) - timedelta(days=1)
    
    if date_range == 'd':
        start = this_day_start
        end = this_day_end
    elif date_range == 'w':
        start = this_week_start
        end = this_week_end
    elif date_range == 'm':
        start = this_month_start
        end = this_month_end
    elif date_range == 'q':
        start = this_quarter_start
        end = this_quarter_end
    else:
        start = this_year_start
        end = this_year_end
    try:
        ws = Department.objects.get(dep_id=workshop)
        if not ws:
            return HttpResponse('{ \"errmsg\": \"不存在此车间。\" }', content_type='text/json')

        wt = WorkType.objects.get(work_type_id=worktype)
        if not wt:
            return HttpResponse('{ \"errmsg\": \"不存在此工种。\" }', content_type='text/json')
        paper_list = TestPapers.objects.values('weixin_open_id')\
            .annotate(paper_count=Count('test_paper_id')).values('weixin_open_id', 'paper_count')\
            .filter(date_time__gte=start, date_time__lte=end, weixin_open_id__in=Members.objects.filter(work_type_id=wt, dep_id=ws))\
            .order_by('-paper_count')[:100]
        ret_list = list()
        for p in paper_list:
            mem_name = Members.objects.get(weixin_open_id=p['weixin_open_id'])
            ret_list.append({ 'name': mem_name.name, 'count': p['paper_count'] })

        return HttpResponse(json.dumps(ret_list), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getMembers(request):
    name = request.POST.get('name', '')
    phone = request.POST.get('phone', '')
    workshop = request.POST.get('workshop', '')
    worktype = request.POST.get('worktype', '')
    id_card = request.POST.get('idcard', '')
    position = request.POST.get('position', '')
    threenew = request.POST.get('three_new', '0')

    try:
        m_workshop = ''
        m_worktype = ''
        m_position = ''

        if workshop != '':
            m_workshop = Department.objects.get(dep_id=int(workshop))
        if m_workshop is None:
            return HttpResponse('{ \"errmsg\": \"车间不存在。\" }', content_type='text/json')

        if worktype != '':
            m_worktype = WorkType.objects.get(work_type_id=int(worktype))
        if m_worktype is None:
            return HttpResponse('{ \"errmsg\": \"工种不存在。\" }', content_type='text/json')
        
        if position != '':
            m_position = Position.objects.get(position_id=int(position))
        if m_position is None:
            return HttpResponse('{ \"errmsg\": \"职名不存在。\" }', content_type='text/json')

        conds = list()
        if name:
            conds.append(Q(name__icontains=name))
        if phone:
            conds.append(Q(phone_number__contains=phone))
        if id_card:
            conds.append(Q(idcard__contains=id_card))
        if workshop != '':
            conds.append(Q(dep_id=m_workshop))
        if worktype != '':
            conds.append(Q(work_type_id=m_worktype))
        if position != '':
            conds.append(Q(position_id=m_position))
        
        conds.append(Q(three_new=(True if threenew == '1' else False)))

        if len(conds) != 0:
            mems = Members.objects.filter(reduce(operator.and_, conds))
        else:
            mems = Members.objects.all()

        ret = list()
        for m in mems:
            ret.append({ 'name': m.name, \
                        'phone': m.phone_number, \
                        'workshop': m.dep_id.dep_name, \
                        'ws_value': m.dep_id.dep_id, \
                        'worktype': m.work_type_id.type_name, \
                        'wt_value': m.work_type_id.work_type_id, \
                        'position': m.position_id.name, \
                        'pos_value': m.position_id.position_id, \
                        'idcard': m.idcard, \
                        'weixin_open_id': m.weixin_open_id, \
                        'verified': ('是' if m.verified else '否'), \
                        'deleted': ('是' if m.deleted else '否'), \
                        'three_new': ('是' if m.three_new  else '否'), \
                        'intro': m.intro })
        return HttpResponse(json.dumps(ret), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getInfoTreeTop(request):
    start = request.POST.get('date_start', None)
    end = request.POST.get('date_end', None)
    ws = request.POST.get('workshop', None)
    wt = request.POST.get('worktype', None)
    pos = request.POST.get('position', None)   

    try:
        conds = list()
        paper_conds = list()
        if start and end and start != '' and end != '':
            start = start + ' 00:00:00'
            end = end + ' 23:59:59'
            paper_conds.append(Q(date_time__gte=start) & Q(date_time__lte=end))

        if ws and ws != '':
            if Department.objects.filter(dep_id=int(ws)).exists():
                conds.append(Q(dep_id=int(ws)))
            else:
                return HttpResponse("{ \"errmsg\": \"车间不存在。\" }", content_type='text/json')

        if wt and wt != '':
            if WorkType.objects.filter(work_type_id=int(wt)).exists():
                conds.append(Q(work_type_id=int(wt)))
            else:
                return HttpResponse("{ \"errmsg\": \"工种不存在。\" }", content_type='text/json')

        if pos and pos != '':
            if Position.objects.filter(position_id=int(pos)).exists():
                conds.append(Q(position_id=int(pos)))
            else:
                return HttpResponse("{ \"errmsg\": \"岗位不存在。\" }", content_type='text/json')

        conds.append(Q(verified=True) & Q(deleted=False) & Q(weixin_open_id__isnull=False) & (~Q(weixin_open_id='')))
        ret = list()
        members = Members.objects.filter(reduce(operator.and_, conds)).values_list('weixin_open_id', flat=True)
        papers = ExamPapers.objects\
            .filter(weixin_open_id__in=members).filter(reduce(operator.and_, paper_conds))\
            .values('paper_id')\
            .annotate(exam_count=Count('exam_paper_id'))\
            .values('paper_id', 'exam_count', 'name', 'passing_score')
            
        for p in papers:
            should_mem_count = p['exam_count']
            examinee_count = ExamPapers.objects.filter(weixin_open_id__in=members, paper_id=p['paper_id'], done=True).count()
            avg_score = ExamPapers.objects.filter(weixin_open_id__in=members, paper_id=p['paper_id'], done=True).aggregate(avg_score=Avg('score'))
            failed_count = ExamPapers.objects.filter(weixin_open_id__in=members, paper_id=p['paper_id'], score__lt=p['passing_score'], done=True).aggregate(failed=Count('exam_paper_id'))
            ret.append({\
                'paper_id': p['paper_id'],\
                'paper_name': p['name'],\
                'examinee_count': examinee_count,\
                'should_mem_count': should_mem_count,\
                'failed_count': failed_count['failed'],\
                'avg_score': avg_score['avg_score']
            })
        
        return HttpResponse(json.dumps(ret), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getMissedDetail(request):
    p_id = request.POST.get('paper_id', None)
    ws_id = request.POST.get('workshop', '')
    wt_id = request.POST.get('worktype', '')
    pos_id = request.POST.get('position', '')

    if not p_id:
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    mem_conds = list()
    if ws_id != '':
        mem_conds.append(Q(dep_id=int(ws_id)))
    if wt_id != '':
        mem_conds.append(Q(work_type_id=int(wt_id)))
    if pos_id != '':
        mem_conds.append(Q(position_id=int(pos_id)))

    mem_conds.append(Q(verified=True) & Q(deleted=False) & Q(weixin_open_id__isnull=False) & (~Q(weixin_open_id='')))
    try:
        ret = list()
        mems = Members.objects.filter(reduce(operator.and_, mem_conds)).order_by('dep_id', 'work_type_id', 'position_id').values_list('weixin_open_id', flat=True)
        papers = ExamPapers.objects.filter(paper_id=p_id, done=False, weixin_open_id__in=mems)

        for p in papers:
            if not Members.objects.filter(weixin_open_id=p.weixin_open_id).exists():
                continue

            mem = Members.objects.get(weixin_open_id=p.weixin_open_id)
            ret.append({
                'name': mem.name,
                'dep_name': mem.dep_id.dep_name,
                'type_name': mem.work_type_id.type_name,
                'pos_name': mem.position_id.name
            })

        return HttpResponse(json.dumps(ret), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getFailedDetail(request):
    p_id = request.POST.get('paper_id', None)
    ws_id = request.POST.get('workshop', '')
    wt_id = request.POST.get('worktype', '')
    pos_id = request.POST.get('position', '')

    if not p_id:
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    mem_conds = list()
    if ws_id != '':
        mem_conds.append(Q(dep_id=int(ws_id)))
    if wt_id != '':
        mem_conds.append(Q(work_type_id=int(wt_id)))
    if pos_id != '':
        mem_conds.append(Q(position_id=int(pos_id)))

    mem_conds.append(Q(verified=True) & Q(deleted=False) & Q(weixin_open_id__isnull=False) & (~Q(weixin_open_id='')))
    try:
        ret = list()
        mems = Members.objects.filter(reduce(operator.and_, mem_conds)).order_by('dep_id', 'work_type_id', 'position_id').values_list('weixin_open_id', flat=True)
        papers = ExamPapers.objects.filter(paper_id=p_id, weixin_open_id__in=mems, done=True)

        for p in papers:
            if (not Members.objects.filter(weixin_open_id=p.weixin_open_id).exists()) or (p.score >= p.passing_score):
                continue

            mem = Members.objects.get(weixin_open_id=p.weixin_open_id)
            ret.append({
                'name': mem.name,
                'dep_name': mem.dep_id.dep_name,
                'type_name': mem.work_type_id.type_name,
                'pos_name': mem.position_id.name,
                'score': p.score
            })

        return HttpResponse(json.dumps(ret), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getScoreDetail(request):
    p_id = request.POST.get('paper_id', None)
    ws_id = request.POST.get('workshop', '')
    wt_id = request.POST.get('worktype', '')
    pos_id = request.POST.get('position', '')

    if not p_id:
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    mem_conds = list()
    if ws_id != '':
        #mem_conds.append(Q(dep_id=int(ws_id)))
        ws_cond = ' and m.dep_id=' + ws_id
    else:
        ws_cond = ''

    if wt_id != '':
        #mem_conds.append(Q(work_type_id=int(wt_id)))
        wt_cond = ' and m.work_type_id=' + wt_id
    else:
        wt_cond = ''

    if pos_id != '':
        #mem_conds.append(Q(position_id=int(pos_id)))
        pos_cond = ' and m.position_id=' + pos_id
    else:
        pos_cond = ''

    #mem_conds.append(Q(verified=True) & Q(deleted=False) & Q(weixin_open_id__isnull=False) & (~Q(weixin_open_id='')))
    try:
        data = list()
        #mems = Members.objects.filter(reduce(operator.and_, mem_conds)).order_by('dep_id', 'work_type_id', 'position_id').values_list('weixin_open_id', flat=True)
        #papers = ExamPapers.objects.filter(paper_id=p_id, weixin_open_id__in=mems)
        
        papers = ExamPapers.objects.raw('select ep.exam_paper_id, ep.done, ep.paper_id, ep.name, ep.weixin_open_id, ep.score from ExamPapers ep inner join Members m on ep.weixin_open_id=m.weixin_open_id and ifnull(m.weixin_open_id, \'\')<>\'\' and m.verified=1 and m.deleted=0 where ep.paper_id=\'' + p_id + '\'' + ws_cond + wt_cond + pos_cond + ' order by m.dep_id, m.work_type_id, m.position_id')
        papers_arr = list(papers)
        if papers is not None and len(papers_arr) > 0:
            paper_name = papers_arr[0].name
        else:
            paper_name = ''

        for p in papers:
            if not Members.objects.filter(weixin_open_id=p.weixin_open_id).exists():
                continue

            mem = Members.objects.get(weixin_open_id=p.weixin_open_id)
            data.append({
                'name': mem.name,
                'dep_name': mem.dep_id.dep_name,
                'type_name': mem.work_type_id.type_name,
                'pos_name': mem.position_id.name,
                'score': p.score,
                'missed': ('是' if p.done == False else '')
            })

        if ws_id != '' and Department.objects.filter(dep_id=int(ws_id)).exists():
            ws_name = Department.objects.get(dep_id=int(ws_id)).dep_name
        else:
            ws_name = ''

        if wt_id != '' and WorkType.objects.filter(work_type_id=int(wt_id)).exists():
            wt_name = WorkType.objects.get(work_type_id=int(wt_id)).type_name
        else:
            wt_name = ''

        if pos_id != '' and Position.objects.filter(position_id=int(pos_id)).exists():
            pos_name = Position.objects.get(work_type_id=int(pos_id)).name
        else:
            pos_name = ''

        ret = {
            'workshop_name': ws_name,
            'worktype_name': wt_name,
            'position_name': pos_name,
            'paper_name': paper_name,
            'data': data
        }
        return HttpResponse(json.dumps(ret), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getPaperTypeOptions(request):
    try:
        pts = PaperTypes.objects.all()

        ret = list()
        for p in pts:
            ret.append({ 'value': p.type_id, 'label': p.name })
        return HttpResponse(json.dumps(ret), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getPapersByTypeAdmin(request):
    t_id = request.POST.get('type_id', '')
    if t_id == '':
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    try:
        typeid = PaperTypes.objects.get(type_id=t_id)
        if not typeid:
            return HttpResponse('{ \"errmsg\": \"题类不存在。\" }', content_type='text/json')

        papers = Papers.objects.filter(type_id=typeid)
        ret = list()
        for p in papers:
            ret.append({ 'key': p.paper_id, 'label': p.paper_name, 'disabled': False })
        return HttpResponse(json.dumps(ret))
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getExamMembers(request):
    workshop = request.POST.get('workshop', '')
    worktype = request.POST.get('worktype', '')
    position = request.POST.get('position', '')
    three_new = request.POST.get('three_new', '0')

    try:
        conds = list()

        if workshop != '':
            conds.append('dep_id in(' + workshop + ')')
        if worktype != '':
            conds.append('work_type_id in(' + worktype + ')')
        if position != '':
            conds.append('position_id in(' + position + ')')

        conds.append('three_new=' + three_new)
        conds.append('verified=1')
        conds.append('deleted=0')
        conds.append('ifnull(weixin_open_id, \'\')<>\'\'')

        mems = Members.objects.extra(where=conds)

        ret = list()
        for m in mems:
            ret.append({ 'name': m.name, \
                        'phone': m.phone_number, \
                        'workshop': m.dep_id.dep_name, \
                        'worktype': m.work_type_id.type_name, \
                        'position': m.position_id.name, \
                        'idcard': m.idcard, \
                        'weixin_open_id': m.weixin_open_id, \
                        'verified': ('是' if m.verified else '否'), \
                        'deleted': ('是' if m.deleted else '否'), \
                        'intro': m.intro })
        return HttpResponse(json.dumps(ret), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def createExams(request):
    ws = request.POST.get('workshops', '')
    wt = request.POST.get('worktypes', '')
    pos = request.POST.get('positions','')
    mems = request.POST.get('members', '')
    paper_ids = request.POST.get('paper_ids', '')
    passing_score = request.POST.get('passing_score', '')
    test_name = request.POST.get('test_name', '')
    test_time = request.POST.get('test_time', '')
    test_ss_count = request.POST.get('test_ss_count', '')
    test_ms_count = request.POST.get('test_ms_count', '')
    test_jm_count = request.POST.get('test_jm_count', '')
    avail_start = request.POST.get('avail_start', '')
    avail_end = request.POST.get('avail_end', '')

    try:
        with connection.cursor() as cursor:
            cursor.callproc('CommitExam', (paper_ids, ws, wt, pos, mems, test_time, passing_score, test_name, test_ss_count, test_ms_count, test_jm_count, str(uuid.uuid1()), avail_start, avail_end))

        return HttpResponse('{ \"errmsg\": \"OK\" }', content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type="text/json")

@csrf_exempt
def uploadQuestionLibraryFile(request):
    paper_type = request.POST.get('paper_type', '')
    paper_name = request.POST.get('paper_name', '')
    passing_score = request.POST.get('passing_score', '')
    test_time = request.POST.get('test_time', '')
    worktypes = request.POST.get('worktypes', '')
    positions = request.POST.get('positions', '')
    workshops = request.POST.get('workshops', '')
    f = request.FILES.get('file')

    if paper_type == '' or paper_name == '' or passing_score == '' or test_time == '' or not f:
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    file_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'upload', f.name).encode('utf-8').decode('utf-8')
    try:
        if os.path.exists(file_path):
            os.remove(file_path)

        with open(file_path, 'wb+') as file_uploaded:
            for chunk in f.chunks():
                file_uploaded.write(chunk)

        p_type = PaperTypes.objects.get(type_id=paper_type)
        new_paper = Papers()
        new_paper.paper_name = paper_name
        new_paper.type_id = p_type
        new_paper.set_date = datetime.now()
        new_paper.passing_score = int(passing_score)
        new_paper.test_time = int(test_time)
        new_paper.work_type_id = WorkType.objects.get(work_type_id=1)
        new_paper.save()

        wb = load_workbook(filename=file_path)
        ws = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(ws[0])
        i = 1
        for row in ws.rows:
            if ws.cell(row=i, column=3).value is None or ws.cell(row=i, column=3).value == '' \
                    or ws.cell(row=i, column=1).value is None or ws.cell(row=i, column=1).value == '' \
                    or ws.cell(row=i, column=2).value is None or ws.cell(row=i, column=2).value == '':
                i += 1
                continue

            new_ques = Questions()
            new_ques.paper_id = new_paper
            new_ques.question_title = ws.cell(row=i, column=3).value
            new_ques.question_sn = ws.cell(row=i, column=1).value
            new_ques.question_type = ws.cell(row=i, column=2).value
            if str(ws.cell(row=i, column=2).value).strip() == '3':
                new_ques.question_answer_texts = '%s|%s' % (ws.cell(row=i, column=4).value, ws.cell(row=i, column=5).value)
            else:
                new_ques.question_answer_texts = '%s|%s|%s|%s' % (ws.cell(row=i, column=4).value, ws.cell(row=i, column=5).value, ws.cell(row=i, column=6).value, ws.cell(row=i, column=7).value)
            new_ques.question_right_answers = ws.cell(row=i, column=8).value
            new_ques.save()
            i += 1
        
        os.remove(file_path)

        wt_arr = worktypes.split(',')
        pos_arr = positions.split(',')
        ws_arr = workshops.split(',')

        n_p_id = new_paper.paper_id

        if worktypes == '':
            wts = WorkType.objects.all()
            for w in wts:
                if PaperWorkTypeRange.objects.select_related("work_type_id").filter(work_type_id=w, paper_id=n_p_id).exists():
                    continue

                new_wt_map = PaperWorkTypeRange()
                new_wt_map.paper_id = new_paper
                new_wt_map.work_type_id = w
                new_wt_map.save()
        else:
            for w in wt_arr:
                if not WorkType.objects.filter(work_type_id=w).exists():
                    continue

                if PaperWorkTypeRange.objects.select_related("work_type_id").filter(work_type_id=w, paper_id=n_p_id).exists():
                    continue

                new_wt_map = PaperWorkTypeRange()
                new_wt_map.paper_id = new_paper
                new_wt_map.work_type_id = WorkType.objects.get(work_type_id=w)
                new_wt_map.save()

        if positions == '':
            poses = Position.objects.all()
            for p in poses:
                if PaperPositionRange.objects.filter(position_id=p, paper_id=n_p_id).exists():
                    continue

                new_pos_map = PaperPositionRange()
                new_pos_map.paper_id = new_paper
                new_pos_map.position_id = p
                new_pos_map.save()
        else:
            for p in pos_arr:
                if not Position.objects.filter(position_id=p).exists():
                    continue

                if PaperPositionRange.objects.select_related("position_id").filter(position_id=p, paper_id=n_p_id).exists():
                    continue

                new_pos_map = PaperPositionRange()
                new_pos_map.paper_id = new_paper
                new_pos_map.position_id = Position.objects.get(position_id=p)
                new_pos_map.save()

        if workshops == '':
            wss = Department.objects.all()
            for ws in wss:
                if PaperDepRange.objects.filter(dep_id=ws, paper_id=n_p_id).exists():
                    continue

                new_ws_map = PaperDepRange()
                new_ws_map.paper_id = new_paper
                new_ws_map.dep_id = ws
                new_ws_map.save()
        else:
            for ws in ws_arr:
                if not Department.objects.filter(dep_id=ws).exists():
                    continue

                if PaperDepRange.objects.select_related("dep_id").filter(dep_id=ws, paper_id=n_p_id).exists():
                    continue

                new_ws_map = PaperDepRange()
                new_ws_map.paper_id = new_paper
                new_ws_map.dep_id = Department.objects.get(dep_id=ws)
                new_ws_map.save()

        return HttpResponse('{ \"errmsg\": \"OK\" }', content_type='text/json')
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        return HttpResponse('{ \"errmsg\": \"%s\", \"lineno\": \"%s\" }' % (str(e), exc_tb.tb_lineno), content_type='text/json')

@csrf_exempt
def getQuestionsInfo(request):
    paper_id = request.POST.get('paper_id', '')
    if paper_id == '':
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')
    
    try:
        if not Papers.objects.filter(paper_id=paper_id).exists():
            return HttpResponse('{ \"errmsg\": \"找不到此题库。\" }', content_type='text/json')

        paper = Papers.objects.get(paper_id=paper_id)

        ques_rows = list()
        questions = Questions.objects.filter(paper_id=paper).order_by('question_sn')
        for q in questions:
            answers_list = q.question_answer_texts.split('|')
            ans_len = len(answers_list)
            if ans_len != 2 and ans_len != 4:
                continue

            ques_rows.append({
                                'question_id': q.question_id,
                                'sn': q.question_sn,
                                'question_type': q.question_type,
                                'question_title': q.question_title,
                                'selection1': answers_list[0],
                                'selection2': answers_list[1],
                                'selection3': ('' if ans_len == 2 else answers_list[2]),
                                'selection4': ('' if ans_len == 2 else answers_list[3]),
                                'question_right_answers': q.question_right_answers
                            })
        return HttpResponse(json.dumps({ 'paper_name': paper.paper_name, 'rows': ques_rows }), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def saveEditedQuestion(request):
    q_type = request.POST.get('question_type', '')
    q_title = request.POST.get('question_title', '')
    q_answer_texts = request.POST.get('question_answer_texts', '')
    q_right = request.POST.get('question_right_answers', '')
    q_id = request.POST.get('question_id', '')
    d = request.POST.get('deleted', '0')

    if (d == '0' and (q_type == '' or q_title == '' or q_answer_texts == '' or q_right == '')) or q_id == '':
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    try:
        if not Questions.objects.filter(question_id=q_id).exists():
            return HttpResponse('{ \"errmsg\": \"未找到此试题。\" }', content_type='text/json')

        q = Questions.objects.get(question_id=q_id)

        if d == 0:
            q.question_type = int(q_type)
            q.question_title = q_title
            q.question_answer_texts = q_answer_texts
            q.question_right_answers = q_right
            q.save()
        else:
            q.delete()
            with connection.cursor() as cursor:
                cursor.callproc('make_questions_sn_smooth', (q.paper_id.paper_id,))

        return HttpResponse('{ \"errmsg\": \"OK\" }', content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def deletePaper(request):
    p_id = request.POST.get('paper_id', '')
    if p_id == '':
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    try:
        if not Papers.objects.filter(paper_id=p_id).exists():
            return HttpResponse('{ \"errmsg\": \"未找到此题库。\" }', content_type='text/json')

        p = Papers.objects.get(paper_id=p_id)

        p.delete()
        return HttpResponse('{ \"errmsg\": \"OK\" }', content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def appendQuestions(request):
    paper_id = request.POST.get('paper_id', '')

    f = request.FILES.get('file')

    if paper_id == '':
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    file_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'upload', f.name)
    try:
        if not Papers.objects.filter(paper_id=paper_id).exists():
            return HttpResponse('{ \"errmsg\": \"未找到此题库。\" }', content_type='text/json')

        paper = Papers.objects.get(paper_id=paper_id)

        if os.path.exists(file_path):
            os.remove(file_path)

        with open(file_path, 'wb+') as file_uploaded:
            for chunk in f.chunks():
                file_uploaded.write(chunk)
        
        max_obj = Questions.objects.aggregate(max_sn=Max('sn'))
        cur_sn = int(max_obj.max_sn) + 1

        wb = load_workbook(filename=file_path)
        ws = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(ws[0])
        i = 1
        for row in ws.rows:
            new_ques = Questions()
            new_ques.paper_id = paper
            new_ques.question_title = ws.cell(row=i, column=3).value
            new_ques.question_sn = cur_sn
            new_ques.question_type = ws.cell(row=i, column=2).value
            if int(ws.cell(row=i, column=2).value) == 3:
                new_ques.question_answer_texts = '%s|%s' % (ws.cell(row=i, column=4).value, ws.cell(row=i, column=5).value)
            else:
                new_ques.question_answer_texts = '%s|%s|%s|%s' % (ws.cell(row=i, column=4).value, ws.cell(row=i, column=5).value, ws.cell(row=i, column=6).value, ws.cell(row=i, column=7).value)
            new_ques.question_right_answers = ws.cell(row=i, column=8).value
            new_ques.save()
            
            cur_sn += 1
            i += 1
        
        os.remove(file_path)

        return HttpResponse('{ \"errmsg\": \"OK\" }', content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def newPaperType(request):
    type_name = request.POST.get('type_name', '')
    if type_name == '':
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    try:
        if PaperTypes.objects.filter(name=type_name).exists():
            return HttpResponse('{ \"errmsg\": \"重名题库已存在。\" }', content_type='text/json')

        new_pt = PaperTypes()
        new_pt.name = type_name
        new_pt.save()
        return HttpResponse('{ \"new_type_id\": \"%s\" }' % new_pt.type_id, content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def deletePaperType(request):
    type_id = request.POST.get('type_id', '')
    if type_name == '':
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    try:
        if not PaperTypes.objects.filter(type_id=type_id).exists():
            return HttpResponse('{ \"errmsg\": \"题库不存在。\" }', content_type='text/json')

        pt = PaperTypes.objects.get(type_id=type_id)

        pt.delete()
        return HttpResponse('{ \"errmsg\": \"OK\" }', content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='')

@csrf_exempt
def modMember(request):
    name = request.POST.get('name', '')
    phone = request.POST.get('phone', '')
    idcard = request.POST.get('idcard', '')
    ws_id = request.POST.get('workshop', '')
    wt_id = request.POST.get('worktype', '')
    pos_id = request.POST.get('position', '')
    deleted = request.POST.get('deleted', '0')
    three_new = request.POST.get('three_new', '0')
    intro = request.POST.get('intro', '')

    if name == '' or ws_id == '' or wt_id == '' or pos_id == '' or idcard == '':
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    try:
        if not Department.objects.filter(dep_id=ws_id).exists():
            return HttpResponse('{ \"errmsg\": \"车间不存在。\" }', content_type='text/json')
        if not WorkType.objects.filter(work_type_id=wt_id).exists():
            return HttpResponse('{ \"errmsg\": \"工种不存在。\" }', content_type='text/json')
        if not Position.objects.filter(position_id=pos_id).exists():
            return HttpResponse('{ \"errmsg\": \"职名不存在。\" }', content_type='text/json')
        if not Members.objects.filter(idcard=idcard).exists():
            return HttpResponse('{ \"errmsg\": \"查无此人。\" }', content_type='text/json')

        ws = Department.objects.get(dep_id=ws_id)
        wt = WorkType.objects.get(work_type_id=wt_id)
        pos = Position.objects.get(position_id=pos_id)

        mem = Members.objects.get(idcard=idcard)
        mem.name = name
        mem.phone_number = phone
        mem.idcard = idcard
        mem.dep_id = ws
        mem.work_type_id =wt
        mem.position_id = pos
        mem.deleted = (False if deleted == '0' else True)
        mem.three_new = (False if three_new == '0' else True)
        mem.intro = intro
        mem.save()

        return HttpResponse('{ \"errmsg\": \"OK\" }', content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def addMember(request):
    name = request.POST.get('name', '')
    phone = request.POST.get('phone', '')
    idcard = request.POST.get('idcard', '')
    ws_id = request.POST.get('workshop', '')
    wt_id = request.POST.get('worktype', '')
    pos_id = request.POST.get('position', '')
    deleted = request.POST.get('deleted', '0')
    three_new = request.POST.get('three_new', '0')
    intro = request.POST.get('intro', '')

    if name == '' or ws_id == '' or wt_id == '' or pos_id == '' or idcard == '':
        return HttpResponse('{ \"errmsg\": \"请求异常。\" }', content_type='text/json')

    try:
        if not Department.objects.filter(dep_id=ws_id).exists():
            return HttpResponse('{ \"errmsg\": \"车间不存在。\" }', content_type='text/json')
        if not WorkType.objects.filter(work_type_id=wt_id).exists():
            return HttpResponse('{ \"errmsg\": \"工种不存在。\" }', content_type='text/json')
        if not Position.objects.filter(position_id=pos_id).exists():
            return HttpResponse('{ \"errmsg\": \"职名不存在。\" }', content_type='text/json')
        if Members.objects.filter(idcard=idcard).exists():
            return HttpResponse('{ \"errmsg\": \"用户已存在。\" }', content_type='text/json')

        ws = Department.objects.get(dep_id=ws_id)
        wt = WorkType.objects.get(work_type_id=wt_id)
        pos = Position.objects.get(position_id=pos_id)
        mem = Members()
        
        mem.name = name
        mem.phone_number = phone
        mem.idcard = idcard
        mem.dep_id = ws
        mem.work_type_id = wt
        mem.position_id = pos
        mem.deleted = (False if deleted == '0' else True)
        mem.three_new = (False if three_new == '0' else True)
        mem.intro = intro
        mem.verified = True
        mem.save()

        return HttpResponse('{ \"errmsg\": \"OK\" }', content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getPaperImportLog(request):
    try:
        logs = PaperImportLog.objects.all().order_by('-import_time')[:10]
        ret = list()
        for l in logs:
            if not Papers.objects.filter(paper_id=l.paper_id).exists():
                continue
            pid = Papers.objects.get(paper_id=l.paper_id)
            ret.append({ 'import_time': l.import_time.strftime('%Y年%m月%d日 %H:%M:%S'), 'name': pid.paper_name })

        return HttpResponse(json.dumps(ret), content_type='text/json')
    except Exception as e:
        return HttpResponse('{ \"errmsg\": \"%s\" }' % str(e), content_type='text/json')

@csrf_exempt
def getDWPConstraint(request):
    wss = request.POST.get('workshops', '')
    wts = request.POST.get('worktypes', '')
    poss = request.POST.get('positions', '')

    try:
        conds = list()
        where_str = '' if wss == '' and wts == '' and poss == '' else ' where '
        if wss != '':
            conds.append('dwt.dep_id in(' + wss + ') and wtp.position_id in(select position_id from DepPosition where dep_id in(' + wss + '))')
        if wts != '':
            conds.append('wtp.work_type_id in(' + wts + ')')
        if poss != '':
            conds.append('wtp.position_id in(' + poss + ')')

        conds_str = ' and '.join(conds)
        
        recs = DepWorkType.objects.raw('''select (@cnt := @cnt + 1) as id,dwt.dep_id, wtp.position_id, wtp.work_type_id 
                                          from DepWorkType dwt 
                                          inner join WorkTypePosition wtp on dwt.work_type_id=wtp.work_type_id cross join (select @cnt := 0) t''' + where_str + conds_str)
        ws_ids = set()
        wt_ids = set()
        ps_ids = set()

        for r in recs:
            ws_ids.add(r.dep_id.dep_id)
            wt_ids.add(r.work_type_id.work_type_id)
            ps_ids.add(r.position_id)

#        if wss != '':
#            ws_ids = set(wss.split(','))
#            if wts == '':
#                dwt = DepWorkType.objects.filter(dep_id__in=wss.split(','))
#                for w in dwt:
#                    wt_ids.add(w.work_type_id)
#            else:
#                for w in wts.split(','):
#                    wt_ids.add(w)
#
#            if poss == '':
#                dp = DepPosition.objects.filter(dep_id__in=wss.split(','))
#                for p in dp:
#                    ps_ids.add(p.position_id)
#            else:
#                for p in poss.split(','):
#                    ps_ids.add(p)
#        else:
#            ws_ids = set([ d.dep_id for d in Department.objects.all() ])
#
#            if wts == '':
#

        if len(ws_ids) != 0:
            workshops = Department.objects.filter(dep_id__in=list(ws_ids))
        else:
            workshops = Department.objects.all()

        if len(wt_ids) != 0:
            worktypes = WorkType.objects.filter(work_type_id__in=list(wt_ids))
        else:
            worktypes = WorkType.objects.all()

        if len(ps_ids) != 0:
            positions = Position.objects.filter(position_id__in=list(ps_ids))
        else:
            positions = Position.objects.all()

        ws_ret = list()
        wt_ret = list()
        ps_ret = list()

        for d in workshops:
            ws_ret.append({ 'value': d.dep_id, 'label': d.dep_name })
        for t in worktypes:
            wt_ret.append({ 'value': t.work_type_id, 'label': t.type_name })
        for p in positions:
            ps_ret.append({ 'value': p.position_id, 'label': p.name })

        return HttpResponse('{ \"workshops\": %s, \"worktypes\": %s, \"positions\": %s }' % (json.dumps(ws_ret), json.dumps(wt_ret), json.dumps(ps_ret)), content_type='text/json')
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        return HttpResponse('{ \"errmsg\": \"%s\", \"lineno\": \"%s\" }' % (str(e), exc_tb.tb_lineno), content_type='text/json')

def dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row)) for row in cursor.fetchall()
    ]
def debugLog(mes):
    log = logging.getLogger('nc_exam_file_debug_logger')
    log.setLevel(logging.DEBUG)
    formatter = logging.Formatter('[%(levelname)s][第%(lineno)s行] - %(message)s');
    fh = logging.FileHandler('/home/carzpurzkey/projects/nc_exam/nc_exam/exam_admin/log.log', mode='a', encoding='utf-8')
    fh.setFormatter(formatter)
    log.addHandler(fh)
    log.debug(mes)
